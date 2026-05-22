import csv
import io
import openpyxl
from django.http import HttpResponse


def export_csv(queryset, field_names, header_map, filename):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    response.write('\ufeff'.encode('utf-8'))
    writer = csv.writer(response)
    writer.writerow([header_map.get(f, f) for f in field_names])
    for obj in queryset:
        row = []
        for f in field_names:
            val = obj
            for attr in f.split('__'):
                val = getattr(val, attr, '') if val else ''
            if val is None:
                val = ''
            row.append(str(val))
        writer.writerow(row)
    return response


def export_xlsx(queryset, field_names, header_map, filename):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = filename[:31]
    ws.append([header_map.get(f, f) for f in field_names])
    for obj in queryset:
        row = []
        for f in field_names:
            val = obj
            for attr in f.split('__'):
                val = getattr(val, attr, '') if val else ''
            if val is None:
                val = ''
            row.append(val)
        ws.append(row)
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    return response


def parse_upload(file, fmt):
    rows = []
    if fmt == 'csv':
        decoded = file.read().decode('utf-8-sig')
        reader = csv.DictReader(io.StringIO(decoded))
        for r in reader:
            rows.append({k.strip(): v.strip() for k, v in r.items()})
    elif fmt == 'xlsx':
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), read_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        headers = [str(c).strip() if c else '' for c in next(rows_iter)]
        for row in rows_iter:
            d = {}
            for i, v in enumerate(row):
                if i < len(headers):
                    d[headers[i]] = str(v).strip() if v is not None else ''
            if any(d.values()):
                rows.append(d)
        wb.close()
    return rows
