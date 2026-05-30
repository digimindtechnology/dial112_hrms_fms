import csv
import json
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from django.contrib.auth.decorators import login_required
from django.db import DatabaseError
from django.db.models import Count
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render, get_object_or_404
from accounts.base import Base
from accounts.helpers.decorators import CheckRole
from accounts.models import Profile
from setup.models import (
    CostCenter,
    CostCenterStatus,
    CostCenterBulkUploadHistory,
    EmployeeType,
    EmployeeCategory,
    EmployeeGroup,
    Designation,
    Grade,
    Language,
    Dialect,
    LeaveType,
    ShiftCategory,
    ViolationType,
    TrainerType,
    FRVType,
    FRVMaintenanceType,
    PoliceStation,
    Holiday,
    HolidayDistrictWise,
)
from masters.models import District
from accounts.helpers.import_export_utils import parse_upload


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def data_setup(request):
    lookup_cards = [{'slug': slug, **cfg} for slug, cfg in SETUP_LOOKUPS.items()]
    return render(request, "setup/data_setup.html", {'setup_lookups': lookup_cards})


SETUP_LOOKUPS = {
    'employee-type': {'model': EmployeeType, 'title': 'Employee Type', 'plural': 'Employee Types', 'icon': 'tabler-users-group'},
    'employee-category': {'model': EmployeeCategory, 'title': 'Employee Category', 'plural': 'Employee Categories', 'icon': 'tabler-category'},
    'employee-group': {'model': EmployeeGroup, 'title': 'Employee Group', 'plural': 'Employee Groups', 'icon': 'tabler-users'},
    'designation': {'model': Designation, 'title': 'Designation', 'plural': 'Designations', 'icon': 'tabler-id-badge-2'},
    'grade': {'model': Grade, 'title': 'Grade', 'plural': 'Grades', 'icon': 'tabler-layers-subtract'},
    'language': {'model': Language, 'title': 'Language', 'plural': 'Languages', 'icon': 'tabler-language'},
    'dialect': {'model': Dialect, 'title': 'Dialect', 'plural': 'Dialects', 'icon': 'tabler-message-language'},
    'leave-type': {'model': LeaveType, 'title': 'Leave Type', 'plural': 'Leave Types', 'icon': 'tabler-calendar-minus'},
    'shift-category': {'model': ShiftCategory, 'title': 'Shift Category', 'plural': 'Shift Categories', 'icon': 'tabler-clock-hour-4', 'has_time': True},
    'violation-type': {'model': ViolationType, 'title': 'Violation Type', 'plural': 'Violation Types', 'icon': 'tabler-alert-triangle'},
    'trainer-type': {'model': TrainerType, 'title': 'Trainer Type', 'plural': 'Trainer Types', 'icon': 'tabler-user-star'},
    'frv-type': {'model': FRVType, 'title': 'FRV Type', 'plural': 'FRV Types', 'icon': 'tabler-ambulance'},
    'frv-maintenance-type': {'model': FRVMaintenanceType, 'title': 'FRV Maintenance Type', 'plural': 'FRV Maintenance Types', 'icon': 'tabler-tools'},
}


def _lookup_config(slug):
    return SETUP_LOOKUPS.get(slug)


def _lookup_tenant(request):
    profile = Profile.objects.filter(user=request.user).select_related('tenantProfile').first()
    return profile.tenantProfile if profile else None


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def setup_lookup_list(request, slug):
    cfg = _lookup_config(slug)
    if not cfg:
        return JsonResponse({'success': False, 'error': 'Unknown setup type'}, status=404)
    try:
        items = cfg['model'].objects.filter(tenantProfile_id=request.tenantID).order_by('name')
    except DatabaseError as e:
        items = []
        db_error = str(e)
    else:
        db_error = ''

    return render(request, "setup/setup_lookup_list.html", {
        'slug': slug,
        'items': items,
        'title': cfg['title'],
        'plural': cfg['plural'],
        'has_time': cfg.get('has_time', False),
        'db_error': db_error,
    })


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def setup_lookup_form(request, slug, pk=0):
    cfg = _lookup_config(slug)
    if not cfg:
        return JsonResponse({'success': False, 'error': 'Unknown setup type'}, status=404)

    model = cfg['model']
    try:
        obj = model.objects.filter(id=pk, tenantProfile_id=request.tenantID).first() if pk else None
    except DatabaseError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)

    if pk and not obj:
        return JsonResponse({'success': False, 'error': f"{cfg['title']} not found"}, status=404)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        description = request.POST.get("description", "").strip()
        is_active = request.POST.get("is_active") == "on"
        if not name:
            return JsonResponse({'success': False, 'error': f"{cfg['title']} name is required"})

        start_time = None
        end_time = None
        is_end_next_day = False
        if cfg.get('has_time'):
            st_value = request.POST.get("start_time", "").strip()
            et_value = request.POST.get("end_time", "").strip()
            is_end_next_day = request.POST.get("is_end_next_day") == "on"
            if st_value:
                try:
                    start_time = datetime.strptime(st_value, "%H:%M").time()
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Invalid start time format (use HH:MM)'})
            if et_value:
                try:
                    end_time = datetime.strptime(et_value, "%H:%M").time()
                except ValueError:
                    return JsonResponse({'success': False, 'error': 'Invalid end time format (use HH:MM)'})

        tenant = _lookup_tenant(request)
        if not tenant:
            return JsonResponse({'success': False, 'error': 'No tenant profile found for your account'})

        try:
            duplicate = model.objects.filter(tenantProfile=tenant, name__iexact=name)
            if obj:
                duplicate = duplicate.exclude(id=obj.id)
            if duplicate.exists():
                return JsonResponse({'success': False, 'error': f"{cfg['title']} already exists"})

            if obj:
                obj.name = name
                obj.description = description or None
                obj.is_active = is_active
                if cfg.get('has_time'):
                    obj.start_time = start_time
                    obj.end_time = end_time
                    obj.is_end_next_day = is_end_next_day
                obj.updated_by = request.user
                obj.save()
            else:
                create_kwargs = {
                    'name': name,
                    'description': description or None,
                    'is_active': is_active,
                    'tenantProfile': tenant,
                    'created_by': request.user,
                }
                if cfg.get('has_time'):
                    create_kwargs['start_time'] = start_time
                    create_kwargs['end_time'] = end_time
                    create_kwargs['is_end_next_day'] = is_end_next_day
                model.objects.create(**create_kwargs)
            return JsonResponse({'success': True})
        except (DatabaseError, Exception) as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return render(request, "setup/partials/_setup_lookup_form.html", {
        'slug': slug,
        'obj': obj,
        'title': cfg['title'],
        'has_time': cfg.get('has_time', False),
    })


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def setup_lookup_delete(request, slug):
    cfg = _lookup_config(slug)
    if not cfg:
        return JsonResponse({'success': False, 'error': 'Unknown setup type'}, status=404)

    pk = request.POST.get("pk")
    try:
        cfg['model'].objects.filter(id=pk, tenantProfile_id=request.tenantID).delete()
        return JsonResponse({'success': True})
    except DatabaseError as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def cost_center_list(request):
    tenant_id = request.tenantID
    items = CostCenter.objects.filter(tenantProfile_id=tenant_id).only(
        'cost_center_id', 'cost_center_name', 'cost_center_code', 'description', 'is_active'
    ).order_by('cost_center_name')
    upload_history = CostCenterBulkUploadHistory.objects.filter(tenantProfile_id=tenant_id).select_related('status', 'uploaded_by')
    return render(request, "setup/cost_center_list.html", {'items': items, 'upload_history': upload_history})


@login_required
def cost_center_add(request):
    if request.method == "POST":
        name = request.POST.get("cost_center_name", "").strip()
        description = request.POST.get("description", "").strip()
        is_active = request.POST.get("is_active") == "on"
        if not name:
            return JsonResponse({'success': False, 'error': 'Cost center name is required'})
        profile = Profile.objects.filter(user=request.user).select_related('tenantProfile').first()
        tenant = profile.tenantProfile if profile else None
        if not tenant:
            return JsonResponse({'success': False, 'error': 'No tenant profile found for your account'})
        try:
            CostCenter.objects.create(
                cost_center_name=name,
                description=description or None,
                is_active=is_active,
                tenantProfile=tenant,
                created_by=request.user,
            )
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return render(request, "setup/partials/_costcenter_form.html", {'obj': None})


@login_required
def cost_center_edit(request, pk):
    obj = CostCenter.objects.filter(cost_center_id=pk).first()
    if not obj:
        return JsonResponse({'success': False, 'error': 'Cost center not found'})
    if request.method == "POST":
        name = request.POST.get("cost_center_name", "").strip()
        description = request.POST.get("description", "").strip()
        is_active = request.POST.get("is_active") == "on"
        if not name:
            return JsonResponse({'success': False, 'error': 'Cost center name is required'})
        try:
            obj.cost_center_name = name
            obj.description = description or None
            obj.is_active = is_active
            obj.updated_by = request.user
            obj.save()
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})
    return render(request, "setup/partials/_costcenter_form.html", {'obj': obj})


@login_required
def cost_center_delete(request):
    pk = request.POST.get("pk")
    CostCenter.objects.filter(cost_center_id=pk).delete()
    return JsonResponse({'success': True})


@login_required
def cost_center_download_sample(request, fmt):
    filename = 'cost_center_sample'
    headers = ['Cost Center Name', 'Description']
    sample_rows = [
        ['IT Infrastructure', 'All IT-related costs'],
        ['Marketing', 'Marketing and advertising expenses'],
        ['Operations', 'Day-to-day operational expenses'],
    ]

    if fmt == 'csv':
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        response.write('\ufeff'.encode('utf-8'))
        writer = csv.writer(response)
        writer.writerow(headers)
        writer.writerows(sample_rows)
        return response

    # Excel (XLSX)
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Cost Centers'

    ws.append(headers)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for row in sample_rows:
        ws.append(row)

    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 45

    output = io.BytesIO()
    wb.save(output)
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


@login_required
def cost_center_upload_detail(request, pk):
    history = get_object_or_404(CostCenterBulkUploadHistory, upload_history_id=pk)
    items = CostCenter.objects.filter(costCenterBulkUpload=history).order_by('cost_center_name')
    return render(request, "setup/partials/_upload_detail.html", {'history': history, 'items': items})


@login_required
def cost_center_download_errors(request, pk):
    history = get_object_or_404(CostCenterBulkUploadHistory, upload_history_id=pk)
    try:
        error_rows = json.loads(history.remarks or '[]')
    except (json.JSONDecodeError, TypeError):
        error_rows = []
    if not error_rows:
        error_rows = [{'Cost Center Name': '', 'Description': '', 'Error': 'No error details available'}]

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="cost_center_errors_{pk}.csv"'
    response.write('\ufeff'.encode('utf-8'))
    writer = csv.writer(response)
    writer.writerow(['Cost Center Name', 'Description', 'Error'])
    for row in error_rows:
        writer.writerow([row.get('Cost Center Name', ''), row.get('Description', ''), row.get('Error', '')])
    return response


@login_required
def cost_center_bulk_upload(request):
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid method'})

    profile = Profile.objects.filter(user=request.user).select_related('tenantProfile').first()
    tenant = profile.tenantProfile if profile else None
    if not tenant:
        return JsonResponse({'success': False, 'error': 'No tenant profile found for your account'})

    file = request.FILES.get('file')
    fmt = request.POST.get('format', 'csv')
    if not file:
        return JsonResponse({'success': False, 'error': 'No file uploaded'})

    try:
        rows = parse_upload(file, fmt)
    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Could not read file. Make sure the file format matches your selection. Detail: {e}'})

    if not rows:
        return JsonResponse({'success': False, 'error': 'File is empty or has no valid data rows. Check the file has correct headers: "Cost Center Name", "Description".'})

    status_pending, _ = CostCenterStatus.objects.get_or_create(
        cost_center_status_name='Pending', defaults={'cost_center_status_css': 'warning'}
    )
    history = CostCenterBulkUploadHistory.objects.create(
        file_name=file.name,
        uploaded_by=request.user,
        total_records=len(rows),
        processed_records=0,
        failed_records=0,
        pending_records=len(rows),
        status=status_pending,
        tenantProfile=tenant,
    )

    created = 0
    updated = 0
    errors = []
    error_rows = []

    for idx, row in enumerate(rows, start=2):
        name = row.get('Cost Center Name', '').strip()
        description = row.get('Description', '').strip()
        if not name:
            errors.append(f"Row {idx}: Cost Center Name is required")
            error_rows.append({'Cost Center Name': name, 'Description': description, 'Error': 'Cost Center Name is required'})
            continue
        try:
            existing = CostCenter.objects.filter(cost_center_name=name, tenantProfile=tenant).first()
            if existing:
                existing.description = description or None
                existing.updated_by = request.user
                existing.costCenterBulkUpload = history
                existing.save()
                updated += 1
            else:
                CostCenter.objects.create(
                    cost_center_name=name,
                    description=description or None,
                    is_active=True,
                    tenantProfile=tenant,
                    created_by=request.user,
                    costCenterBulkUpload=history,
                )
                created += 1
        except Exception as e:
            errors.append(f"Row {idx}: {e}")
            error_rows.append({'Cost Center Name': name, 'Description': description, 'Error': str(e)})

    processed = created + updated
    failed = len(errors)
    pending = len(rows) - processed - failed
    status_name = 'Completed' if failed == 0 else 'Completed with Errors'
    status_css = 'success' if failed == 0 else 'warning'
    final_status, _ = CostCenterStatus.objects.get_or_create(
        cost_center_status_name=status_name, defaults={'cost_center_status_css': status_css}
    )
    history.total_records = len(rows)
    history.processed_records = processed
    history.failed_records = failed
    history.pending_records = max(0, pending)
    history.status = final_status
    history.remarks = json.dumps(error_rows) if error_rows else ''
    history.save(update_fields=['total_records', 'processed_records', 'failed_records', 'pending_records', 'status', 'remarks'])

    msg = f"Bulk upload complete: {created} created, {updated} updated"
    if errors:
        msg += f", {len(errors)} errors"

    return JsonResponse({'success': True, 'message': msg, 'created': created, 'updated': updated, 'errors': len(errors)})


# ─── Police Station ───────────────────────────────────────────────────────────

@login_required
@CheckRole(Base.Group.SetupGroup.value)
def police_station_list(request):
    items = PoliceStation.objects.filter(tenantProfile_id=request.tenantID).select_related('district').order_by('police_station_name')
    districts = District.objects.filter(is_active=True).order_by('district_name')
    return render(request, "setup/police_station_list.html", {'items': items, 'districts': districts})


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def police_station_form(request, pk=0):
    obj = PoliceStation.objects.filter(police_station_id=pk, tenantProfile_id=request.tenantID).select_related('district').first() if pk else None
    if pk and not obj:
        return JsonResponse({'success': False, 'error': 'Police station not found'}, status=404)

    districts = District.objects.filter(is_active=True).order_by('district_name')

    if request.method == "POST":
        name = request.POST.get("police_station_name", "").strip()
        district_id = request.POST.get("district_id", "").strip()
        station_code = request.POST.get("station_code", "").strip()
        phone_number = request.POST.get("phone_number", "").strip()
        email = request.POST.get("email", "").strip()
        address = request.POST.get("address", "").strip()
        pincode = request.POST.get("pincode", "").strip()
        latitude = request.POST.get("latitude", "").strip()
        longitude = request.POST.get("longitude", "").strip()
        established_date = request.POST.get("established_date", "").strip()
        is_active = request.POST.get("is_active") == "on"

        if not name:
            return JsonResponse({'success': False, 'error': 'Police station name is required'})
        if not district_id:
            return JsonResponse({'success': False, 'error': 'District is required'})

        district = District.objects.filter(district_id=district_id).first()
        if not district:
            return JsonResponse({'success': False, 'error': 'Selected district does not exist'})

        profile = Profile.objects.filter(user=request.user).select_related('tenantProfile').first()
        tenant = profile.tenantProfile if profile else None
        if not tenant:
            return JsonResponse({'success': False, 'error': 'No tenant profile found for your account'})

        lat = None
        lon = None
        if latitude:
            try:
                lat = Decimal(latitude)
            except InvalidOperation:
                return JsonResponse({'success': False, 'error': 'Invalid latitude value'})
        if longitude:
            try:
                lon = Decimal(longitude)
            except InvalidOperation:
                return JsonResponse({'success': False, 'error': 'Invalid longitude value'})

        est_date = None
        if established_date:
            try:
                est_date = datetime.strptime(established_date, "%Y-%m-%d").date()
            except ValueError:
                return JsonResponse({'success': False, 'error': 'Invalid established date'})

        try:
            if obj:
                obj.police_station_name = name
                obj.district = district
                obj.station_code = station_code or None
                obj.phone_number = phone_number or None
                obj.email = email or None
                obj.address = address or None
                obj.pincode = pincode or None
                obj.latitude = lat
                obj.longitude = lon
                obj.established_date = est_date
                obj.is_active = is_active
                obj.updated_by = request.user
                obj.save()
            else:
                PoliceStation.objects.create(
                    police_station_name=name,
                    district=district,
                    station_code=station_code or None,
                    phone_number=phone_number or None,
                    email=email or None,
                    address=address or None,
                    pincode=pincode or None,
                    latitude=lat,
                    longitude=lon,
                    established_date=est_date,
                    is_active=is_active,
                    tenantProfile=tenant,
                    created_by=request.user,
                )
            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    return render(request, "setup/partials/_police_station_form.html", {'obj': obj, 'districts': districts})


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def police_station_delete(request):
    pk = request.POST.get("pk")
    PoliceStation.objects.filter(police_station_id=pk, tenantProfile_id=request.tenantID).delete()
    return JsonResponse({'success': True})


# ─── Holiday ──────────────────────────────────────────────────────────────────

@login_required
@CheckRole(Base.Group.SetupGroup.value)
def holiday_list(request):
    items = (Holiday.objects
             .filter(tenantProfile_id=request.tenantID)
             .prefetch_related('holiday_HolidayDistrictWise__district')
             .order_by('holiday_date'))
    return render(request, "setup/holiday_list.html", {'items': items})


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def holiday_form(request, pk=0):
    obj = Holiday.objects.filter(holiday_id=pk, tenantProfile_id=request.tenantID).first() if pk else None
    if pk and not obj:
        return JsonResponse({'success': False, 'error': 'Holiday not found'}, status=404)

    all_districts = District.objects.filter(is_active=True).order_by('district_name')

    if request.method == "POST":
        name = request.POST.get("holiday_name", "").strip()
        holiday_date = request.POST.get("holiday_date", "").strip()
        description = request.POST.get("description", "").strip()
        is_national_holiday = request.POST.get("is_national_holiday") == "on"
        is_active = request.POST.get("is_active") == "on"

        if not name:
            return JsonResponse({'success': False, 'error': 'Holiday name is required'})
        if not holiday_date:
            return JsonResponse({'success': False, 'error': 'Holiday date is required'})

        try:
            h_date = datetime.strptime(holiday_date, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse({'success': False, 'error': 'Invalid date format'})

        if not request.tenantID:
            return JsonResponse({'success': False, 'error': 'No tenant profile found.'})

        try:
            if obj:
                obj.holiday_name = name
                obj.holiday_date = h_date
                obj.description = description or None
                obj.is_national_holiday = is_national_holiday
                obj.is_active = is_active
                obj.updated_by = request.user
                obj.save()
                holiday = obj
            else:
                holiday = Holiday.objects.create(
                    holiday_name=name,
                    holiday_date=h_date,
                    description=description or None,
                    is_national_holiday=is_national_holiday,
                    is_active=is_active,
                    tenantProfile_id=request.tenantID,
                    created_by=request.user,
                )

            # Sync district assignments
            selected_ids = set()
            for x in request.POST.getlist("district_ids"):
                try:
                    selected_ids.add(int(x))
                except (ValueError, TypeError):
                    pass
            existing_ids = set(
                HolidayDistrictWise.objects.filter(holiday=holiday).values_list('district_id', flat=True)
            )
            HolidayDistrictWise.objects.filter(holiday=holiday, district_id__in=existing_ids - selected_ids).delete()
            for district in District.objects.filter(district_id__in=selected_ids - existing_ids):
                HolidayDistrictWise.objects.create(
                    holiday=holiday, district=district, tenantProfile_id=request.tenantID
                )

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)})

    assigned_ids = list(
        HolidayDistrictWise.objects.filter(holiday=obj).values_list('district_id', flat=True)
    ) if obj else []

    return render(request, "setup/partials/_holiday_form.html", {
        'obj': obj,
        'all_districts': all_districts,
        'assigned_ids': assigned_ids,
    })


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def holiday_delete(request):
    pk = request.POST.get("pk")
    Holiday.objects.filter(holiday_id=pk, tenantProfile_id=request.tenantID).delete()
    return JsonResponse({'success': True})


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def holiday_district_list(request, pk):
    holiday = get_object_or_404(Holiday, holiday_id=pk, tenantProfile_id=request.tenantID)
    assigned = HolidayDistrictWise.objects.filter(holiday=holiday).select_related('district')
    assigned_ids = list(assigned.values_list('district_id', flat=True))
    all_districts = District.objects.filter(is_active=True).order_by('district_name')
    return render(request, "setup/partials/_holiday_district_list.html", {
        'holiday': holiday,
        'assigned': assigned,
        'all_districts': all_districts,
        'assigned_ids': assigned_ids,
    })


@login_required
@CheckRole(Base.Group.SetupGroup.value)
def holiday_district_save(request, pk):
    """Full sync: replaces current district assignments with the submitted selection."""
    if request.method != "POST":
        return JsonResponse({'success': False, 'error': 'Invalid method'})
    holiday = get_object_or_404(Holiday, holiday_id=pk, tenantProfile_id=request.tenantID)

    raw_ids = request.POST.getlist("district_ids")
    selected_ids = set()
    for x in raw_ids:
        try:
            selected_ids.add(int(x))
        except (ValueError, TypeError):
            pass

    profile = Profile.objects.filter(user=request.user).select_related('tenantProfile').first()
    tenant = profile.tenantProfile if profile else None
    if not tenant:
        return JsonResponse({'success': False, 'error': 'No tenant profile found'})

    existing_ids = set(
        HolidayDistrictWise.objects.filter(holiday=holiday).values_list('district_id', flat=True)
    )
    to_add = selected_ids - existing_ids
    to_remove = existing_ids - selected_ids

    try:
        HolidayDistrictWise.objects.filter(holiday=holiday, district_id__in=to_remove).delete()
        for district in District.objects.filter(district_id__in=to_add):
            HolidayDistrictWise.objects.create(holiday=holiday, district=district, tenantProfile=tenant)
        return JsonResponse({'success': True, 'added': len(to_add), 'removed': len(to_remove)})
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})
